import os
from unittest import TestCase

import openpyxl
from astronverse.datatable.openpyxl import OpenpyxlWrapper

test_excel_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../test.xlsx"))


print(f"Test Excel Path: {test_excel_path}")


class TestOpenpyxl(TestCase):
    
    def test_openpyxl_read_cell(self):
        pyxl = openpyxl.load_workbook(test_excel_path)
        sheet = pyxl.active
        value = sheet.cell(row=11, column=1).value
        print(value)
    
    def test_openpyxl_write_cell(self):
        pyxl = openpyxl.load_workbook(test_excel_path)
        sheet = pyxl.active
        sheet.cell(row=1, column=1, value="Test Value")
        pyxl.save(test_excel_path)
        
    def test_read_cell(self):
        wrapper = OpenpyxlWrapper(file_path=test_excel_path)
        value = wrapper.read_cell(row=1, col=1)
        print(value)
        
    def test_read_row(self):
        wrapper = OpenpyxlWrapper(file_path=test_excel_path)
        values = wrapper.read_row(row_index=1)
        print(values)
        
    def test_read_column(self):
        wrapper = OpenpyxlWrapper(file_path=test_excel_path)
        values = wrapper.read_column(col_index=1)
        print(values)
        
    def test_read_area(self):
        wrapper = OpenpyxlWrapper(file_path=test_excel_path)
        values = wrapper.read_range("A1:C3")
        print(values)
    
    def test_read_all(self):
        wrapper = OpenpyxlWrapper(file_path=test_excel_path)
        values = wrapper.read_effective_area()
        print(values)
        
    def test_get_max_row(self):
        wrapper = OpenpyxlWrapper(file_path=test_excel_path)
        max_row = wrapper.get_max_row()
        print(max_row)
    
    def test_get_max_column(self):
        wrapper = OpenpyxlWrapper(file_path=test_excel_path)
        max_col = wrapper.get_max_column()
        print(max_col)
    
    def test_write_cell(self):
        wrapper = OpenpyxlWrapper(file_path=test_excel_path)
        wrapper.write_cell(row=1, col=1, value="11")
        
    def test_insert_cell(self):
        wrapper = OpenpyxlWrapper(file_path=test_excel_path)
        # wrapper.insert_cell(row=2, col=2, value="22", shift="right")
        
    def test_write_cell_formula(self):
        # 测试写入单元格公式
        wrapper = OpenpyxlWrapper(file_path=test_excel_path)
        wrapper.write_cell(row=12, col=1, value="=SUM(A10:A11)")
        wrapper.save()
        value = wrapper.read_cell(row=12, col=1)
        print(value)

