import csv
import json
import os

import openpyxl
from astronverse.datatable.error import *
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class OpenpyxlWrapper:
    def __init__(self, file_path: str, sheet_name=None):
        """
        Initializes the Excel wrapper.

        Args:
            file_path (str): Path to the Excel file.
            sheet_name (str, optional): The name of the sheet to activate.
                                      If None, the default active sheet is used.
                                      If the sheet does not exist, it will be created.
        """
        self.file_path = file_path
        if os.path.exists(file_path):
            self.workbook = openpyxl.load_workbook(file_path)
        else:
            self.workbook = Workbook()

        if sheet_name:
            if sheet_name in self.workbook.sheetnames:
                self.sheet = self.workbook[sheet_name]
            else:
                self.sheet = self.workbook.create_sheet(title=sheet_name)
        else:
            self.sheet = self.workbook.active

    def save(self, path: str = None):
        """
        Saves the workbook.

        Args:
            path (str, optional): The path to save the file. If None, overwrites the original file.
        """
        save_path = path or self.file_path
        try:
            self.workbook.save(save_path)
        except PermissionError:
            raise DATAFRAME_EXPECTION(WRITE_PERMISSION_DENIED_ERROR_FORMAT.format(save_path), "写入Excel文件失败")
        except Exception as e:
            raise DATAFRAME_EXPECTION(WRITE_DATA_ERROR_FORMAT.format(save_path, str(e)), "写入Excel文件失败")

    def close(self):
        """
        Closes the workbook.
        """
        self.workbook.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.save()
        self.close()

    def switch_sheet(self, sheet_name: str):
        """
        Switches the active sheet.

        Args:
            sheet_name (str): The name of the sheet to activate. If it does not exist, it will be created.
        """
        if sheet_name in self.workbook.sheetnames:
            self.sheet = self.workbook[sheet_name]
        else:
            self.sheet = self.workbook.create_sheet(title=sheet_name)

    def add_sheet(self, title: str = None, index: int = None) -> Worksheet:
        """
        Adds a new sheet.

        Args:
            title (str, optional): The title of the new sheet.
            index (int, optional): The position to insert the new sheet.

        Returns:
            Worksheet: The newly created sheet.
        """
        new_sheet = self.workbook.create_sheet(title=title, index=index)
        return new_sheet

    def delete_sheet(self, sheet_name: str):
        """
        Deletes a sheet.

        Args:
            sheet_name (str): The name of the sheet to delete.
        """
        if sheet_name in self.workbook.sheetnames:
            sheet_to_delete = self.workbook[sheet_name]
            self.workbook.remove(sheet_to_delete)

    def copy_sheet(self, source_sheet_name: str, new_sheet_name: str) -> Worksheet:
        """
        Copies a sheet.

        Args:
            source_sheet_name (str): The name of the sheet to copy.
            new_sheet_name (str): The name of the new sheet.

        Returns:
            Worksheet: The newly created sheet.
        """
        source_sheet = self.workbook[source_sheet_name]
        new_sheet = self.workbook.copy_worksheet(source_sheet)
        new_sheet.title = new_sheet_name
        return new_sheet

    def get_sheet_names(self) -> list:
        """
        Gets the names of all sheets.

        Returns:
            list: A list of sheet names.
        """
        return self.workbook.sheetnames

    def rename_sheet(self, old_name: str, new_name: str):
        """
        Renames a sheet.

        Args:
            old_name (str): The current name of the sheet.
            new_name (str): The new name for the sheet.
        """
        if old_name in self.workbook.sheetnames:
            sheet = self.workbook[old_name]
            sheet.title = new_name

    def read_cell(self, row: int, col: int):
        """
        Reads the value of a specific cell.

        Args:
            row (int): The row index (1-based).
            col (int): The column index (1-based).

        Returns:
            The value of the cell.
        """
        return self.sheet.cell(row=row, column=col).value

    def read_row(self, row_index: int) -> list:
        """
        Reads a full row.

        Args:
            row_index (int): The row index (1-based).

        Returns:
            list: A list of cell values in the row.
        """
        return [cell.value for cell in self.sheet[row_index]]

    def read_column(self, col_name: str = None, col_index: int = None) -> list:
        """
        Reads a full column by name or index.

        Args:
            col_name (str, optional): The name of the column (e.g., 'A').
            col_index (int, optional): The index of the column (1-based).

        Returns:
            list: A list of cell values in the column.

        Raises:
            ValueError: If neither col_name nor col_index is provided.
        """
        if col_name:
            return [cell.value for cell in self.sheet[col_name]]
        elif col_index:
            return [cell.value for cell in self.sheet[get_column_letter(col_index)]]
        else:
            raise ValueError("Either column name or column index must be provided.")

    def read_range(self, range_str: str) -> list:
        """
        Reads a range of cells.

        Args:
            range_str (str): The range string (e.g., 'A1:B2').

        Returns:
            list: A 2D list of cell values in the range.
        """
        return [[cell.value for cell in row] for row in self.sheet[range_str]]

    def read_effective_area(self) -> list:
        """
        Reads the effective area of the sheet (all non-empty cells).

        Returns:
            list: A 2D list of cell values in the effective area.
        """
        max_row = self.sheet.max_row
        max_col = self.sheet.max_column
        return [[self.sheet.cell(row=r, column=c).value for c in range(1, max_col + 1)] for r in range(1, max_row + 1)]

    def get_max_row(self) -> int:
        """
        Gets the maximum row index with data.

        Returns:
            int: The maximum row index.
        """
        return self.sheet.max_row

    def get_max_column(self) -> int:
        """
        Gets the maximum column index with data.

        Returns:
            int: The maximum column index.
        """
        return self.sheet.max_column

    def write_cell(self, row: int, col: int, value):
        """
        Writes a value to a specific cell.

        Args:
            row (int): The row index (1-based).
            col (int): The column index (1-based).
            value: The value to write.
        """
        self.sheet.cell(row=row, column=col, value=value)

    def write_row(self, row_index: int, data: list, start_col: int = 1):
        """
        Writes a list of data to a row.

        Args:
            row_index (int): The row index (1-based).
            data (list): The list of data to write.
            start_col (int): The starting column index (1-based).
        """
        print(f"Writing data to row {row_index} starting at column {start_col}: {data}")
        for i, value in enumerate(data):
            self.sheet.cell(row=row_index, column=start_col + i, value=value)

    def append_row(self, data: list):
        """
        Appends a row of data to the end of the sheet.

        Args:
            data (list): The list of data to append.
        """
        self.sheet.append(data)

    def write_column(self, col_name: str = None, col_index: int = None, data: list = None, start_row: int = 1):
        """
        Writes a list of data to a column.

        Args:
            col_name (str, optional): The name of the column (e.g., 'A').
            col_index (int, optional): The index of the column (1-based).
            data (list): The list of data to write.
            start_row (int): The starting row index (1-based).

        Raises:
            ValueError: If neither col_name nor col_index is provided.
        """
        if not col_name and not col_index:
            raise ValueError("Either column name or column index must be provided.")

        col = col_index or openpyxl.utils.column_index_from_string(col_name)

        for i, value in enumerate(data):
            self.sheet.cell(row=start_row + i, column=col, value=value)

    def write_range(self, range_str: str, data: list):
        """
        Writes a 2D list of data to a specified range.

        Args:
            range_str (str): The range string (e.g., 'A1:B2').
            data (list): A 2D list of data to write.
        """
        from openpyxl.utils import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(range_str)

        for r_idx, row_data in enumerate(data):
            for c_idx, cell_value in enumerate(row_data):
                self.sheet.cell(row=min_row + r_idx, column=min_col + c_idx, value=cell_value)

    def fill_data_table_by_import_file(
        self, import_file_path: str, delimiter: str = ",", include_header: bool = True, sheet_name=None
    ):
        ext = os.path.splitext(import_file_path)[1].lower()
        self.sheet.delete_rows(1, self.sheet.max_row)
        if ext == ".csv":
            try:
                with open(import_file_path, newline="", encoding="utf-8") as csvfile:
                    reader = csv.reader(csvfile, delimiter=delimiter)
                    start_row = 1
                    for row_data in reader:
                        self.write_row(row_index=start_row, data=row_data)
                        start_row += 1
            except UnicodeDecodeError:
                with open(import_file_path, newline="", encoding="gbk") as csvfile:
                    reader = csv.reader(csvfile, delimiter=delimiter)
                    start_row = 1
                    for row_data in reader:
                        self.write_row(row_index=start_row, data=row_data)
                        start_row += 1
        if ext in [".xlsx", ".xls"]:
            wb = openpyxl.load_workbook(import_file_path)
            wb_sheet = wb[sheet_name] if sheet_name else wb.active
            max_row = wb_sheet.max_row
            max_col = wb_sheet.max_column
            data = [
                [wb_sheet.cell(row=r, column=c).value for c in range(1, max_col + 1)] for r in range(1, max_row + 1)
            ]
            start_row = 1
            if not include_header:
                data = data[1:]
            for r_idx, row_data in enumerate(data):
                self.write_row(row_index=start_row, data=row_data)
                start_row += 1
            wb.close()

    def insert_cells(self, row: int, col: int, amount: int = 1):
        """
        Inserts blank cells at a specific position.

        Args:
            row (int): The row index (1-based) where to insert cells.
            col (int): The column index (1-based) where to insert cells.
            amount (int): The number of cells to insert.
        """
        self.sheet.insert_cols(idx=col, amount=amount)
        self.sheet.insert_rows(idx=row, amount=amount)

    def insert_rows(self, idx: int, amount: int = 1):
        """
        Inserts blank rows.

        Args:
            idx (int): The row index (1-based) where to insert rows.
            amount (int): The number of rows to insert.
        """
        self.sheet.insert_rows(idx=idx, amount=amount)

    def insert_cols(self, idx: int, amount: int = 1):
        """
        Inserts blank columns.

        Args:
            idx (int): The column index (1-based) where to insert columns.
            amount (int): The number of columns to insert.
        """
        self.sheet.insert_cols(idx=idx, amount=amount)

    def copy_paste_range(self, source_range_str: str, dest_start_cell_str: str):
        """
        Copies a range of cells and pastes it to a new location.

        Args:
            source_range_str (str): The source range (e.g., 'A1:B2').
            dest_start_cell_str (str): The top-left cell of the destination (e.g., 'C1').
        """
        source_range = self.sheet[source_range_str]
        dest_start_cell = self.sheet[dest_start_cell_str]

        dest_start_row = dest_start_cell.row
        dest_start_col = dest_start_cell.column

        for i, row in enumerate(source_range):
            for j, cell in enumerate(row):
                self.sheet.cell(row=dest_start_row + i, column=dest_start_col + j, value=cell.value)

    def delete_cell(self, row: int, col: int, move_direction: str = "up"):
        """
        Deletes a cell and shifts other cells.

        Args:
            row (int): The row index (1-based) of the cell to delete.
            col (int): The column index (1-based) of the cell to delete.
            move_direction (str): The direction to shift cells ('up' or 'left' or 'no' ).
        """
        if move_direction == "up":
            for r in range(row, self.sheet.max_row):
                self.sheet.cell(row=r, column=col).value = self.sheet.cell(row=r + 1, column=col).value
            self.sheet.cell(row=self.sheet.max_row, column=col).value = None
        elif move_direction == "left":
            for c in range(col, self.sheet.max_column):
                self.sheet.cell(row=row, column=c).value = self.sheet.cell(row=row, column=c + 1).value
            self.sheet.cell(row=row, column=self.sheet.max_column).value = None
        else:
            self.sheet.cell(row=row, column=col).value = None

    def delete_rows(self, idx: int, amount: int = 1):
        """
        Deletes rows and shifts cells up.

        Args:
            idx (int): The row index (1-based) from where to delete.
            amount (int): The number of rows to delete.
        """
        self.sheet.delete_rows(idx=idx, amount=amount)

    def delete_cols(self, idx: int, amount: int = 1):
        """
        Deletes columns and shifts cells left.

        Args:
            idx (int): The column index (1-based) from where to delete.
            amount (int): The number of columns to delete.
        """
        self.sheet.delete_cols(idx=idx, amount=amount)

    def empty_row(self, row_index: int):
        """
        Empties the content of a specific row.

        Args:
            row_index (int): The row index (1-based) to empty.
        """
        for cell in self.sheet[row_index]:
            cell.value = None

    def empty_column(self, col_name: str = None, col_index: int = None):
        """
        Empties the content of a specific column.

        Args:
            col_name (str, optional): The name of the column (e.g., 'A').
            col_index (int, optional): The index of the column (1-based).

        Raises:
            ValueError: If neither col_name nor col_index is provided.
        """
        if not col_name and not col_index:
            raise ValueError("Either column name or column index must be provided.")

        col = col_index or openpyxl.utils.column_index_from_string(col_name)

        for row in self.sheet.iter_rows(min_col=col, max_col=col):
            for cell in row:
                cell.value = None

    def clear_range(self, range_str: str):
        """
        Clears the content of a range of cells.

        Args:
            range_str (str): The range to clear (e.g., 'A1:B10').
        """
        for row in self.sheet[range_str]:
            for cell in row:
                cell.value = None

    def sort_range(self, range_str: str, sort_column_index: int, reverse: bool = False):
        """
        Sorts a range based on a specific column.

        Args:
            range_str (str): The range to sort (e.g., 'A1:C10').
            sort_column_index (int): The column index within the range to sort by (0-based).
            reverse (bool): If True, sorts in descending order.
        """
        data = self.read_range(range_str)

        # Extract the top-left cell coordinates to write back the sorted data
        from openpyxl.utils import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(range_str)

        # Sort the data
        sorted_data = sorted(data, key=lambda x: x[sort_column_index], reverse=reverse)

        # Write the sorted data back to the sheet
        for r_idx, row_data in enumerate(sorted_data):
            for c_idx, cell_value in enumerate(row_data):
                self.sheet.cell(row=min_row + r_idx, column=min_col + c_idx, value=cell_value)

    def find_and_replace(self, find_value, replace_value, range_str: str = None):
        """
        Finds and replaces values within a specified range or the entire sheet.

        Args:
            find_value: The value to find.
            replace_value: The value to replace with.
            range_str (str, optional): The range to search in (e.g., 'A1:D10').
                                      If None, searches the entire sheet.
        """
        if range_str:
            target_range = self.sheet[range_str]
        else:
            target_range = self.sheet.iter_rows()

        for row in target_range:
            for cell in row:
                if cell.value == find_value:
                    cell.value = replace_value

    def import_from_csv(self, csv_file_path: str, delimiter=","):
        """
        Imports data from a CSV file into the current sheet.

        Args:
            csv_file_path (str): Path to the CSV file.
            delimiter (str): The delimiter used in the CSV file.
        """
        with open(csv_file_path, newline="", encoding="utf-8") as csvfile:
            reader = csv.reader(csvfile, delimiter=delimiter)
            for row_data in reader:
                self.sheet.append(row_data)

    def export_to_csv(self, csv_file_path: str, include_header: bool = True, delimiter=","):
        """
        Exports the sheet's data to a CSV file.

        Args:
            csv_file_path (str): Path to save the CSV file.
            include_header (bool): Whether to include the column names as the first row.
            delimiter (str): The delimiter to use in the CSV file.
        """
        with open(csv_file_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile, delimiter=delimiter)

            if include_header:
                header = [cell.value for cell in self.sheet[1]]
                writer.writerow(header)

            for row in self.sheet.iter_rows(min_row=2 if include_header else 1):
                writer.writerow([cell.value for cell in row])

    def import_from_json(self, json_file_path: str, include_header: bool = True):
        """
        Imports data from a JSON file (list of lists or list of dicts).

        Args:
            json_file_path (str): Path to the JSON file.
            include_header (bool): If the JSON is a list of dicts, this will write the keys as a header row.
        """
        with open(json_file_path, encoding="utf-8") as jsonfile:
            data = json.load(jsonfile)
            if not data:
                return

            if isinstance(data[0], dict):
                if include_header:
                    headers = list(data[0].keys())
                    self.append_row(headers)
                for item in data:
                    self.append_row(list(item.values()))
            elif isinstance(data[0], list):
                for row_data in data:
                    self.append_row(row_data)

    def export_to_json(self, json_file_path: str, use_header: bool = True):
        """
        Exports the sheet's data to a JSON file.

        Args:
            json_file_path (str): Path to save the JSON file.
            use_header (bool): If True, exports as a list of dictionaries using the first row as keys.
                               Otherwise, exports as a list of lists.
        """
        data = []
        if use_header:
            header = [cell.value for cell in self.sheet[1]]
            for row in self.sheet.iter_rows(min_row=2):
                row_data = {header[i]: cell.value for i, cell in enumerate(row)}
                data.append(row_data)
        else:
            for row in self.sheet.iter_rows():
                data.append([cell.value for cell in row])

        with open(json_file_path, "w", encoding="utf-8") as jsonfile:
            json.dump(data, jsonfile, indent=4)

    def export_to_file(self, file_path: str):
        """
        Exports the workbook to a specified file.

        Args:
            file_path (str): The path to save the file.
        """
        self.save(file_path)

    def get_row_count(self) -> int:
        """
        Gets the number of rows in the sheet.

        Returns:
            int: The maximum row count.
        """
        return self.sheet.max_row

    def get_column_count(self) -> int:
        """
        Gets the number of columns in the sheet.

        Returns:
            int: The maximum column count.
        """
        return self.sheet.max_column

    def get_column_name(self, col_index: int) -> str:
        """
        Gets the column name (e.g., 'A') from its index.

        Args:
            col_index (int): The column index (1-based).

        Returns:
            str: The column name.
        """
        return get_column_letter(col_index)

    def set_column_name(self, col_index: int, name: str):
        """
        Sets the name of a column (writes to the first row).

        Args:
            col_index (int): The column index (1-based).
            name (str): The name to set.
        """
        self.sheet.cell(row=1, column=col_index, value=name)

    def insert_formula(self, row: int, col: int, formula: str):
        """
        Inserts a formula into a cell.

        Args:
            row (int): The row index (1-based).
            col (int): The column index (1-based).
            formula (str): The formula to insert (e.g., '=SUM(A1:A5)').
        """
        self.sheet.cell(row=row, column=col, value=formula)

    def iter_rows(self, min_row=None, max_row=None, min_col=None, max_col=None, values_only=False):
        """
        Provides an iterator over rows in a given range.

        Args:
            min_row (int, optional): The starting row index.
            max_row (int, optional): The ending row index.
            min_col (int, optional): The starting column index.
            max_col (int, optional): The ending column index.
            values_only (bool): If True, yields only cell values.

        Returns:
            generator: A generator for rows.
        """
        return self.sheet.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=values_only
        )

    def iter_cols(self, min_row=None, max_row=None, min_col=None, max_col=None, values_only=False):
        """
        Provides an iterator over columns in a given range.

        Args:
            min_row (int, optional): The starting row index.
            max_row (int, optional): The ending row index.
            min_col (int, optional): The starting column index.
            max_col (int, optional): The ending column index.
            values_only (bool): If True, yields only cell values.

        Returns:
            generator: A generator for columns.
        """
        return self.sheet.iter_cols(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=values_only
        )

    def filter_data(self, condition_callback, range_str: str = None) -> list:
        """
        Filters data based on a callback function.

        Args:
            condition_callback (function): A function that takes a row (as a tuple of cells) and returns True if the row should be included.
            range_str (str, optional): The range to filter. If None, the entire sheet is used.

        Returns:
            list: A list of rows that meet the condition.
        """
        filtered_rows = []

        if range_str:
            rows = self.sheet[range_str]
        else:
            rows = self.sheet.iter_rows()

        for row in rows:
            if condition_callback(row):
                filtered_rows.append([cell.value for cell in row])

        return filtered_rows
