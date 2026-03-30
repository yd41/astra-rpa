from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from astronverse.openapi.nlp.text_moderation import nlp_text_moderation


@patch("astronverse.openapi.nlp.text_moderation.GatewayClient.post")
def test_nlp_text_moderation_exports_excel_when_blocked(mock_post, tmp_path):
    mock_post.return_value = {
        "code": "000000",
        "desc": "success",
        "sid": "sid",
        "data": {
            "request_id": "req-1",
            "result": {
                "suggest": "block",
                "detail": {
                    "content": "test content",
                    "category_list": [
                        {
                            "confidence": 98,
                            "category": "pornDetection",
                            "suggest": "block",
                            "category_description": "色情",
                            "word_list": ["违规词"],
                            "word_infos": [{"word": "违规词", "positions": [1, 3]}],
                        }
                    ],
                },
            },
        },
    }

    result = nlp_text_moderation(
        input_type="text",
        content="test content",
        categories=["pornDetection"],
        is_save=True,
        dst_file=str(tmp_path),
        dst_file_name="text_moderation",
    )

    saved_file = Path(result["saved_file"])

    assert saved_file.exists()
    assert result["data"]["data"]["result"]["suggest"] == "block"
    mock_post.assert_called_once_with(
        "/nlp/text-moderation",
        {
            "content": "test content",
            "is_match_all": 1,
            "categories": ["pornDetection"],
        },
    )

    workbook = load_workbook(saved_file)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0] == ("置信度", "敏感分类", "内容建议结果", "识别类型", "敏感词列表", "敏感词", "敏感词位置下标信息")
    assert rows[1] == ("98", "色情", "block", "色情", "违规词", "违规词", "1,3")


@patch("astronverse.openapi.nlp.text_moderation.GatewayClient.post")
def test_nlp_text_moderation_skips_excel_when_passed(mock_post, tmp_path, capsys):
    mock_post.return_value = {
        "code": "000000",
        "desc": "success",
        "sid": "sid",
        "data": {
            "request_id": "req-1",
            "result": {
                "suggest": "pass",
                "detail": {
                    "content": "正常文本",
                    "category_list": [],
                },
            },
        },
    }

    result = nlp_text_moderation(
        input_type="text",
        content="正常文本",
        is_save=True,
        dst_file=str(tmp_path),
        dst_file_name="text_moderation",
    )

    captured = capsys.readouterr()

    assert result["saved_file"] == ""
    assert result["data"]["data"]["result"]["suggest"] == "pass"
    assert "审核通过，无审核建议需导出" in captured.out
