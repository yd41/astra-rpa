from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from astronverse.openapi.nlp.text_correction import nlp_text_correction


@patch("astronverse.openapi.nlp.text_correction.GatewayClient.post")
def test_nlp_text_correction_exports_excel_and_corrected_text(mock_post, tmp_path):
    mock_post.return_value = {
        "header": {"code": 0, "message": "success", "sid": "sid"},
        "result": {
            "ret": 0,
            "word": [(0, "abc", "xyz", "word")],
            "char": [],
        },
    }

    error_dir = tmp_path / "excel"
    corrected_dir = tmp_path / "txt"

    result = nlp_text_correction(
        input_type="text",
        text="abc def",
        error_dst_file=str(error_dir),
        error_dst_file_name="text_correction",
        export_corrected_doc=True,
        corrected_dst_file=str(corrected_dir),
        corrected_dst_file_name="corrected_doc",
    )

    error_file = Path(result["error_detail_file"])
    corrected_file = Path(result["corrected_file"])

    assert result["data"]["word"][0][2] == "xyz"
    assert error_file.exists()
    assert corrected_file.exists()
    assert corrected_file.read_text(encoding="utf-8") == "xyz def"

    workbook = load_workbook(error_file)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0] == ("错误位置", "错误文本", "纠正文本", "错误类型")
    assert rows[1] == ("0", "abc", "xyz", "词语")


@patch("astronverse.openapi.nlp.text_correction.GatewayClient.post")
def test_nlp_text_correction_skips_excel_when_no_errors(mock_post, tmp_path, capsys):
    mock_post.return_value = {
        "header": {"code": 0, "message": "success", "sid": "sid"},
        "result": {
            "ret": 0,
            "word": [],
            "char": [],
        },
    }

    result = nlp_text_correction(
        input_type="text",
        text="正常文本",
        error_dst_file=str(tmp_path),
        error_dst_file_name="text_correction",
    )

    captured = capsys.readouterr()

    assert result["error_detail_file"] == ""
    assert result["data"]["word"] == []
    assert "无错误，无需导出文档" in captured.out


@patch("astronverse.openapi.nlp.text_correction.GatewayClient.post")
def test_nlp_text_correction_allows_skipping_error_export(mock_post, tmp_path):
    mock_post.return_value = {
        "header": {"code": 0, "message": "success", "sid": "sid"},
        "result": {
            "ret": 0,
            "word": [(0, "abc", "xyz", "word")],
            "char": [],
        },
    }

    corrected_dir = tmp_path / "txt"

    result = nlp_text_correction(
        input_type="text",
        text="abc def",
        is_save=False,
        export_corrected_doc=True,
        corrected_dst_file=str(corrected_dir),
        corrected_dst_file_name="corrected_doc",
    )

    corrected_file = Path(result["corrected_file"])

    assert result["error_detail_file"] == ""
    assert corrected_file.exists()
    assert corrected_file.read_text(encoding="utf-8") == "xyz def"
