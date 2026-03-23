import os
from collections.abc import Generator
from typing import Any

from astronverse.scheduler.logger import logger
from openpyxl import Workbook, load_workbook


class ExcelService:
    """Excel 文件读写服务"""

    def __init__(self, resource_dir: str):
        """
        初始化 Excel 服务

        Args:
            resource_dir: 工程资源目录
        """
        self.resource_dir = resource_dir

    def get_file_path(self, filename: str) -> str:
        """
        获取 Excel 文件的完整路径

        Args:
            filename: 文件名（不含扩展名）

        Returns:
            完整的文件路径
        """
        if not filename.endswith(".xlsx"):
            filename = f"{filename}.xlsx"
        return os.path.join(self.resource_dir, filename)

    def file_exists(self, filename: str) -> bool:
        """
        检查文件是否存在

        Args:
            filename: 文件名

        Returns:
            文件是否存在
        """
        file_path = self.get_file_path(filename)
        return os.path.exists(file_path)

    def create_file(self, filename: str) -> str:
        """
        创建空白 Excel 文件

        Args:
            filename: 文件名

        Returns:
            创建的文件路径
        """
        file_path = self.get_file_path(filename)

        # 确保目录存在
        dir_path = os.path.dirname(file_path)
        if dir_path and not os.path.exists(dir_path):
            os.makedirs(dir_path, exist_ok=True)

        # 创建空白工作簿
        wb = Workbook()
        wb.save(file_path)
        wb.close()

        logger.info(f"Created Excel file: {file_path}")
        return file_path

    def read_file_stream(self, filename: str) -> Generator[dict]:
        """
        流式读取 Excel 文件，逐行返回数据

        Args:
            filename: 文件名

        Yields:
            每行数据的字典，格式为 {"sheet": str, "row": int, "data": list}
        """
        file_path = self.get_file_path(filename)

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        wb = load_workbook(file_path, read_only=True, data_only=True)

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # 发送 sheet 开始事件
                yield {
                    "type": "sheet_start",
                    "sheet": sheet_name,
                    "max_row": ws.max_row or 0,
                    "max_column": ws.max_column or 0,
                }

                # 逐行读取数据
                row_num = 0
                for row in ws.iter_rows(values_only=True):
                    row_num += 1
                    # 将单元格值转换为可序列化的格式
                    row_data = [self._serialize_cell_value(cell) for cell in row]
                    yield {
                        "type": "row",
                        "sheet": sheet_name,
                        "row": row_num,
                        "data": row_data,
                    }

                # 发送 sheet 结束事件
                yield {
                    "type": "sheet_end",
                    "sheet": sheet_name,
                }

        finally:
            wb.close()

        # 发送完成事件
        yield {
            "type": "complete",
            "filename": filename,
        }

    def read_file(self, filename: str) -> dict:
        """
        一次性读取整个 Excel 文件

        Args:
            filename: 文件名

        Returns:
            包含所有数据的字典
        """
        file_path = self.get_file_path(filename)

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        wb = load_workbook(file_path, read_only=True, data_only=False)

        try:
            result = {
                "filename": filename,
                "sheets": [],
                "active_sheet": wb.active.title if wb.active else None,
            }

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_data = {
                    "name": sheet_name,
                    "max_row": ws.max_row or 0,
                    "max_column": ws.max_column or 0,
                    "data": [],
                }

                for row in ws.iter_rows(values_only=True):
                    row_data = [self._serialize_cell_value(cell) for cell in row]
                    sheet_data["data"].append(row_data)

                result["sheets"].append(sheet_data)

            return result

        finally:
            wb.close()

    def write_file(self, filename: str, data: dict) -> None:
        """
        写入数据到 Excel 文件

        Args:
            filename: 文件名
            data: 要写入的数据，格式为 {"sheets": [{"name": str, "data": list[list]}]}
        """
        file_path = self.get_file_path(filename)

        # 确保目录存在
        dir_path = os.path.dirname(file_path)
        if dir_path and not os.path.exists(dir_path):
            os.makedirs(dir_path, exist_ok=True)

        wb = Workbook()

        # 删除默认创建的 sheet
        if wb.active:
            wb.remove(wb.active)

        sheets = data.get("sheets", [])
        if not sheets:
            # 如果没有数据，至少创建一个空 sheet
            wb.create_sheet("Sheet1")
        else:
            for sheet_info in sheets:
                sheet_name = sheet_info.get("name", "Sheet1")
                ws = wb.create_sheet(sheet_name)

                sheet_data = sheet_info.get("data", [])
                for row_idx, row_data in enumerate(sheet_data, start=1):
                    for col_idx, cell_value in enumerate(row_data, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=cell_value)

        # 设置活动 sheet
        active_sheet = data.get("active_sheet")
        if active_sheet and active_sheet in wb.sheetnames:
            wb.active = wb[active_sheet]
        elif wb.sheetnames:
            wb.active = wb[wb.sheetnames[0]]

        wb.save(file_path)
        wb.close()

        logger.info(f"Saved Excel file: {file_path}")

    def update_cells(self, filename: str, updates: list[dict]) -> None:
        """
        更新指定单元格的值

        Args:
            filename: 文件名
            updates: 更新列表，每项格式为 {"sheet": str, "row": int, "col": int, "value": any}
        """
        file_path = self.get_file_path(filename)

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        wb = load_workbook(file_path)

        try:
            for update in updates:
                sheet_name = update.get("sheet")
                row = update.get("row") + 1
                col = update.get("col") + 1
                value = update.get("value")

                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(sheet_name)
                else:
                    ws = wb[sheet_name]

                ws.cell(row=row, column=col, value=value)

            wb.save(file_path)
            logger.info(f"Updated {len(updates)} cells in: {file_path}")

        finally:
            wb.close()

    def delete_file(self, filename: str) -> bool:
        """
        删除 Excel 文件

        Args:
            filename: 文件名

        Returns:
            是否删除成功
        """
        file_path = self.get_file_path(filename)

        if os.path.exists(file_path):
            os.remove(file_path)
            logger.info(f"Deleted Excel file: {file_path}")
            return True

        return False

    @staticmethod
    def _serialize_cell_value(value: Any) -> Any:
        """
        将单元格值转换为可 JSON 序列化的格式

        Args:
            value: 单元格原始值

        Returns:
            序列化后的值
        """
        if value is None:
            return None
        elif isinstance(value, (str, int, float, bool)):
            return value
        else:
            # 其他类型转换为字符串
            return str(value)
